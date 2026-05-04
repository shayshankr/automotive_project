"""
publish_results.py
==================
Adds the "Churn_Results" sheet to Task .xlsx — a color-coded, ranked,
final published churn risk output that the examiner can see immediately
on opening the file.

Also adds a "Summary" banner showing counts and revenue-at-risk.

Run: python solution_code/publish_results.py
  (or called automatically by run_solution.py)
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
import config as C

# ── helpers ───────────────────────────────────────────────────────────────────
def paint(cell, value, bg, fg="000000", bold=False, size=10,
          h="center", v="center", wrap=True):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    s = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=s, right=s, top=s, bottom=s)

def merge_paint(ws, r, c1, c2, value, bg, fg="FFFFFF",
                bold=True, size=11, h="center"):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(r, c1)
    paint(c, value, bg, fg, bold=bold, size=size, h=h)
    ms = Side(style="medium", color="555555")
    c.border = Border(left=ms, right=ms, top=ms, bottom=ms)

def col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rh(ws, row, h):
    ws.row_dimensions[row].height = h

# Risk tier colours
TIER_BG = {"High": "FFD7D7", "Medium": "FFF2CC", "Low": "E2EFDA"}
TIER_FG = {"High": "C00000", "Medium": "C55A11", "Low": "375623"}
TIER_ICON = {"High": "🔴 HIGH", "Medium": "🟡 MEDIUM", "Low": "🟢 LOW"}


def build_results_sheet():
    full = C.OUTPUT_FILE
    wb   = openpyxl.load_workbook(full)

    # ── Load & join data ──────────────────────────────────────────────────────
    def load(sheet):
        df = pd.read_excel(full, sheet_name=sheet, header=1)
        return df.iloc[1:].reset_index(drop=True)   # skip schema-notes row

    cust    = load("Customer")
    scores  = load("Churn_Scores")
    feats   = load("Churn_Features")
    vehicle = load("Vehicle")
    sales   = load("Sales")

    # sales → vehicle model/variant via vin
    veh_info = vehicle[["vin","model","variant","color","fuel_type_id"]].drop_duplicates("vin")
    sal_veh  = sales[["cust_id","vin"]].merge(veh_info, on="vin", how="left")

    # master join
    df = scores.merge(cust[["cust_id","name_masked","mobile_masked","city","state","gender","occ_id"]],
                      on="cust_id", how="left")
    df = df.merge(feats[["cust_id","vehicle_id","vehicle_age_years","total_service_visits",
                          "days_since_last_service","days_to_policy_expiry",
                          "avg_bill_amount","rfm_total","churn_label"]],
                  on="cust_id", how="left")
    df = df.merge(sal_veh[["cust_id","model","variant"]], on="cust_id", how="left")

    # Sort: High first, then Medium, then Low; within each group by prob desc
    tier_order = {"High": 0, "Medium": 1, "Low": 2}
    df["_tier_ord"] = df["risk_tier"].map(tier_order)
    df = df.sort_values(["_tier_ord","churn_probability"], ascending=[True, False])
    df = df.reset_index(drop=True)
    df["rank"] = df.index + 1

    # Revenue at risk estimate (avg INR 15,000 per customer/year)
    REV_PER_CUST = 15000
    high_count   = int((df.risk_tier == "High").sum())
    med_count    = int((df.risk_tier == "Medium").sum())
    low_count    = int((df.risk_tier == "Low").sum())
    total        = len(df)
    rev_at_risk  = high_count * REV_PER_CUST

    # ── Build / replace sheet ─────────────────────────────────────────────────
    if "Churn_Results" in wb.sheetnames:
        del wb["Churn_Results"]
    ws = wb.create_sheet("Churn_Results", 0)   # put it first
    wb.move_sheet("Churn_Results", offset=0)
    ws.sheet_properties.tabColor = "C00000"

    # Column widths
    col_w(ws, [6, 10, 18, 16, 14, 12, 12, 14, 18, 18, 20, 22, 22, 42])
    # Rank, CustID, Name_masked, City, Model, Variant, Prob, RiskTier,
    # DaysSinceSvc, DaysToExpiry, TopDriver1, TopDriver2, TopDriver3, Action

    NCOLS = 14

    # ── Row 1: Main title ─────────────────────────────────────────────────────
    merge_paint(ws, 1, 1, NCOLS,
        "CHURN RISK RESULTS  —  Automotive After-Sales Customer Retention Dashboard",
        "1F4E79", "FFFFFF", bold=True, size=14)
    rh(ws, 1, 34)

    # ── Row 2: Summary stats ──────────────────────────────────────────────────
    rh(ws, 2, 22)
    summary_cells = [
        (1,  2,  f"Total Customers: {total}",             "404040"),
        (3,  4,  f"🔴 HIGH RISK:  {high_count} ({high_count*100//total}%)", "C00000"),
        (5,  6,  f"🟡 MEDIUM RISK: {med_count} ({med_count*100//total}%)", "C55A11"),
        (7,  8,  f"🟢 LOW RISK:   {low_count} ({low_count*100//total}%)",  "375623"),
        (9,  11, f"Est. Annual Revenue at Risk (High only): INR {rev_at_risk:,}", "7B0000"),
        (12, 14, "Sorted: High → Medium → Low | Descending Probability",    "555555"),
    ]
    for c1, c2, text, fg in summary_cells:
        ws.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
        c = ws.cell(2, c1)
        paint(c, text, "F2F2F2", fg, bold=True, size=10, h="center")

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    headers = ["Rank","Cust ID","Name (Masked)","City","Model","Variant",
               "Churn Prob","Risk Tier","Days Since\nLast Service",
               "Days to Policy\nExpiry","Top Driver 1","Top Driver 2",
               "Top Driver 3","Recommended Action"]
    rh(ws, 3, 32)
    for ci, h in enumerate(headers, 1):
        paint(ws.cell(3, ci), h, "1F4E79", "FFFFFF", bold=True, size=10,
              h="center", wrap=True)

    # ── Rows 4+: Data ─────────────────────────────────────────────────────────
    for row_idx, rec in df.iterrows():
        excel_row = row_idx + 4
        tier  = str(rec.get("risk_tier","Low"))
        bg    = TIER_BG.get(tier, "FFFFFF")
        fg    = TIER_FG.get(tier, "000000")
        prob  = rec.get("churn_probability", 0)
        try:
            prob_f = float(prob)
        except:
            prob_f = 0.0

        rh(ws, excel_row, 20)

        row_data = [
            (str(int(rec["rank"])),              "F7F7F7", "333333", True,  "center"),
            (str(rec.get("cust_id","")),          "EEF4FB", "1F4E79", True,  "center"),
            (str(rec.get("name_masked","")),       "FFFFFF", "000000", False, "left"),
            (str(rec.get("city","")),              "FFFFFF", "333333", False, "center"),
            (str(rec.get("model","")),             "FFFFFF", "000000", True,  "center"),
            (str(rec.get("variant","")),           "FFFFFF", "333333", False, "center"),
            (f"{prob_f:.4f}",                      bg,       fg,       True,  "center"),
            (TIER_ICON.get(tier, tier),            bg,       fg,       True,  "center"),
        ]

        # Days since last service — highlight if > 365
        dss = rec.get("days_since_last_service", 0)
        try:
            dss_i = int(float(dss))
        except:
            dss_i = 0
        dss_bg = "FFD7D7" if dss_i > 365 else ("FFF2CC" if dss_i > 180 else "E2EFDA")
        row_data.append((str(dss_i),             dss_bg, TIER_FG.get("High" if dss_i>365 else "Medium" if dss_i>180 else "Low","333333"), False, "center"))

        # Days to policy expiry — highlight if negative
        dte = rec.get("days_to_policy_expiry", 0)
        try:
            dte_i = int(float(dte))
        except:
            dte_i = 0
        dte_bg = "FFD7D7" if dte_i < 0 else ("FFF2CC" if dte_i < 30 else "E2EFDA")
        dte_lbl = f"{dte_i} (EXPIRED)" if dte_i < 0 else str(dte_i)
        row_data.append((dte_lbl, dte_bg, TIER_FG.get("High" if dte_i<0 else "Medium" if dte_i<30 else "Low","333333"), False, "center"))

        row_data += [
            (str(rec.get("top_driver_1","")), "FAFAFA", "555555", False, "left"),
            (str(rec.get("top_driver_2","")), "FAFAFA", "555555", False, "left"),
            (str(rec.get("top_driver_3","")), "FAFAFA", "555555", False, "left"),
            (str(rec.get("recommended_action","")), bg, fg, False, "left"),
        ]

        for ci, (val, cbg, cfg, cbold, ch) in enumerate(row_data, 1):
            paint(ws.cell(excel_row, ci), val, cbg, cfg, bold=cbold,
                  size=9, h=ch, wrap=True)

    ws.freeze_panes = "A4"

    # ── Legend row at bottom ──────────────────────────────────────────────────
    last_row = total + 5
    rh(ws, last_row, 18)
    merge_paint(ws, last_row, 1, NCOLS,
        "LEGEND:  🔴 HIGH = No service > 365 days AND insurance expired  |  "
        "🟡 MEDIUM = Service gap 181-365 days OR policy expiring < 30 days  |  "
        "🟢 LOW = Service within 180 days AND active policy",
        "404040", "FFFFFF", bold=False, size=9)

    # Try saving to the main file; if locked (e.g. open in Excel),
    # save to a standalone Churn_Results_Published.xlsx instead.
    import os as _os
    alt_path = _os.path.join(_os.path.dirname(full), "Churn_Results_Published.xlsx")
    try:
        wb.save(full)
        saved_to = full
    except PermissionError:
        # Main file is open — save to standalone file
        wb2 = openpyxl.Workbook()
        # copy just the Churn_Results sheet into new workbook
        wb2.remove(wb2.active)
        from openpyxl import load_workbook as _lw
        # re-build into a fresh workbook
        wb.save(alt_path)
        saved_to = alt_path

    print(f"\n  Churn_Results sheet saved to: {saved_to}")
    print(f"  Churn_Results       OK  ({total} customers ranked)")
    print(f"    High:   {high_count}  ({high_count*100//total}%)")
    print(f"    Medium: {med_count}  ({med_count*100//total}%)")
    print(f"    Low:    {low_count}  ({low_count*100//total}%)")
    print(f"    Revenue at risk: INR {rev_at_risk:,}")
    if saved_to == alt_path:
        print(f"\n  NOTE: Task .xlsx was open in another app.")
        print(f"  The results were saved to: {alt_path}")
        print(f"  Close Task .xlsx and re-run to merge into main file.")


if __name__ == "__main__":
    build_results_sheet()
    print("\nChurn_Results sheet published.")
