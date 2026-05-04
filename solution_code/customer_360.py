"""
customer_360.py
===============
Pulls a complete 360-degree view of any single customer across all sheets
in Task .xlsx — masked PII, raw PII (from vault), vehicle, sales, service
history, insurance history, churn features, churn score, and SHAP drivers.

Always saves the result to <CUST_ID>.xlsx (e.g. CUST0030.xlsx) with one
sheet per data category, plus a printable console summary.

USAGE
-----
  python solution_code/customer_360.py CUST0030
  python solution_code/customer_360.py CUST0013 --show-pii
"""

import os, sys, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config as C


# ── helpers ──────────────────────────────────────────────────────────────────
def load(sheet):
    df = pd.read_excel(C.OUTPUT_FILE, sheet_name=sheet, header=1)
    return df.iloc[1:].reset_index(drop=True)


def section(title, char="="):
    print("\n" + char * 90)
    print(f"  {title}")
    print(char * 90)


def show_df(df, label):
    if df is None or df.empty:
        print(f"  [no {label} records found]")
        return
    print(df.to_string(index=False))


# ── Excel styling ────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_COLORS = {
    "Summary":         ("1F4E79", "FFFFFF"),
    "Customer":        ("375623", "FFFFFF"),
    "PII_Vault":       ("C00000", "FFFFFF"),
    "Vehicle":         ("BF8F00", "FFFFFF"),
    "Sales":           ("2E75B6", "FFFFFF"),
    "Service":         ("2E75B6", "FFFFFF"),
    "Insurance":       ("2E75B6", "FFFFFF"),
    "Churn_Features":  ("7030A0", "FFFFFF"),
    "Churn_Score":     ("7030A0", "FFFFFF"),
}


def write_sheet(wb, name, df, title):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    bg, fg = SHEET_COLORS.get(name, ("404040", "FFFFFF"))
    ws.sheet_properties.tabColor = bg

    # Title row
    ncols = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1)
    c.value = title
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", color=fg, bold=True, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    if df is None or df.empty:
        ws.cell(2, 1).value = f"[no records found]"
        ws.cell(2, 1).font = Font(italic=True, color="888888")
        return

    # Header row
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(2, ci)
        cell.value = str(col)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Arial", color=fg, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[2].height = 24

    # Data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        bg_row = "FFFFFF" if ri % 2 else "F7F9FC"
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(ri, ci)
            v = row[col]
            cell.value = "" if pd.isna(v) else v
            cell.fill = PatternFill("solid", fgColor=bg_row)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = BORDER

    # Column widths
    for ci, col in enumerate(df.columns, 1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].fillna("").astype(str)]
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A3"


# ── core ─────────────────────────────────────────────────────────────────────
def customer_360(cust_id, show_pii=False):
    cust_id = cust_id.strip().upper()

    section(f"CUSTOMER 360° VIEW  —  {cust_id}", "#")

    cust = load("Customer")
    rec = cust[cust["cust_id"] == cust_id]
    if rec.empty:
        print(f"\n  Customer {cust_id} not found.\n")
        return None

    # Console output
    section("1.  CUSTOMER MASTER  (masked)")
    show_df(rec.T.rename(columns={rec.index[0]: "value"}), "customer")

    pii_df = pd.DataFrame()
    section("2.  PII VAULT  (raw)" + ("" if show_pii else "  [hidden]"))
    if show_pii:
        pii = load("PII_Vault")
        pii_df = pii[pii["cust_id"] == cust_id]
        show_df(pii_df.T.rename(columns={pii_df.index[0]: "value"}) if not pii_df.empty else pii_df, "PII")
    else:
        print("  Use --show-pii flag (authorised access only).")

    section("3.  VEHICLE(S) OWNED")
    sales = load("Sales")
    cv = sales[sales["cust_id"] == cust_id]
    veh_df = pd.DataFrame()
    if not cv.empty:
        vehicle = load("Vehicle")
        veh_df = vehicle[vehicle["vin"].isin(cv["vin"])]
        show_df(veh_df, "vehicle")
    else:
        print("  [no vehicles linked]")

    section("4.  SALES TRANSACTIONS")
    show_df(cv, "sales")

    section("5.  SERVICE HISTORY")
    svc = load("Service")
    s_df = svc[svc["cust_id"] == cust_id]
    if "service_date" in s_df.columns:
        s_df = s_df.sort_values("service_date", ascending=False)
    show_df(s_df, "service")

    section("6.  INSURANCE HISTORY")
    ins = load("Insurance")
    i_df = ins[ins["cust_id"] == cust_id]
    show_df(i_df, "insurance")

    section("7.  CHURN FEATURES  (engineered)")
    feats = load("Churn_Features")
    f_df = feats[feats["cust_id"] == cust_id]
    if not f_df.empty:
        show_df(f_df.T.rename(columns={f_df.index[0]: "value"}), "features")

    section("8.  CHURN SCORE  &  RECOMMENDED ACTION")
    scores = load("Churn_Scores")
    sc_df = scores[scores["cust_id"] == cust_id]
    tier, prob = "?", "?"
    if not sc_df.empty:
        show_df(sc_df.T.rename(columns={sc_df.index[0]: "value"}), "score")
        tier = sc_df.iloc[0].get("risk_tier", "?")
        prob = sc_df.iloc[0].get("churn_probability", "?")
        icon = {"High": "[HIGH]", "Medium": "[MED]", "Low": "[LOW]"}.get(str(tier), "")
        print(f"\n  Verdict:  {icon}  {tier}-RISK   (probability = {prob})")

    # ── Export to <CUST_ID>.xlsx ─────────────────────────────────────────────
    out_path = os.path.join(C.BASE_DIR, f"{cust_id}.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet first
    summary = pd.DataFrame({
        "field": [
            "Customer ID", "Name (masked)", "City", "State",
            "Risk Tier", "Churn Probability",
            "Days Since Last Service", "Days to Policy Expiry",
            "Total Service Visits", "Policy Renewals",
            "Top Driver 1", "Top Driver 2", "Top Driver 3",
            "Recommended Action",
        ],
        "value": [
            cust_id,
            rec.iloc[0].get("name_masked", ""),
            rec.iloc[0].get("city", ""),
            rec.iloc[0].get("state", ""),
            tier,
            prob,
            f_df.iloc[0].get("days_since_last_service", "") if not f_df.empty else "",
            f_df.iloc[0].get("days_to_policy_expiry", "")   if not f_df.empty else "",
            f_df.iloc[0].get("total_service_visits", "")    if not f_df.empty else "",
            f_df.iloc[0].get("num_policy_renewals", "")     if not f_df.empty else "",
            sc_df.iloc[0].get("top_driver_1", "") if not sc_df.empty else "",
            sc_df.iloc[0].get("top_driver_2", "") if not sc_df.empty else "",
            sc_df.iloc[0].get("top_driver_3", "") if not sc_df.empty else "",
            sc_df.iloc[0].get("recommended_action", "") if not sc_df.empty else "",
        ],
    })
    write_sheet(wb, "Summary",        summary, f"360° SUMMARY — {cust_id}")
    write_sheet(wb, "Customer",       rec,     "CUSTOMER MASTER (Masked)")
    if show_pii and not pii_df.empty:
        write_sheet(wb, "PII_Vault",  pii_df,  "PII VAULT (Raw — Authorised Access)")
    write_sheet(wb, "Vehicle",        veh_df,  "VEHICLE(S) OWNED")
    write_sheet(wb, "Sales",          cv,      "SALES TRANSACTIONS")
    write_sheet(wb, "Service",        s_df,    "SERVICE HISTORY")
    write_sheet(wb, "Insurance",      i_df,    "INSURANCE HISTORY")
    write_sheet(wb, "Churn_Features", f_df,    "CHURN FEATURES (Engineered)")
    write_sheet(wb, "Churn_Score",    sc_df,   "CHURN SCORE & RECOMMENDED ACTION")

    try:
        wb.save(out_path)
        print(f"\n  Report saved to: {out_path}")
    except PermissionError:
        alt_path = os.path.join(C.BASE_DIR, f"{cust_id}_new.xlsx")
        wb.save(alt_path)
        print(f"\n  {out_path} was open — saved to: {alt_path}")

    section("END OF 360° REPORT", "#")
    return out_path


def main():
    ap = argparse.ArgumentParser(description="360° lookup for one customer.")
    ap.add_argument("cust_id", help="Customer ID, e.g. CUST0030")
    ap.add_argument("--show-pii", action="store_true",
                    help="Unmask raw PII (authorised use only)")
    args = ap.parse_args()
    customer_360(args.cust_id, show_pii=args.show_pii)


if __name__ == "__main__":
    main()
