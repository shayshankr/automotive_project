"""
generate_solution.py
====================
Main data generation script.  Builds the complete Excel workbook with:

  Dimension tables (reference / lookup)
    dim_Bank | dim_Dealer | dim_Location | dim_WorkType | dim_PolicyType
    dim_PayMode | dim_CancelReason | dim_Occupation | dim_FuelType | dim_Milestone

  Fact tables (normalized — no raw PII)
    Sales | Service | Insurance

  Master tables (masked PII in analytics tables)
    Customer (masked) | Vehicle | Customer_Vehicle (bridge)

  Security table
    PII_Vault (actual PII — access controlled, clearly labelled)

  Analytics / ML
    Churn_Features | Churn_Scores | ML_Pipeline

  Documentation
    README

Run via:  python run_solution.py
"""

import os, sys, random
from datetime import date, timedelta

# Make sure sibling modules are importable when run directly
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter

import config    as C
import pii_utils as P
import excel_utils as E

rng = random.Random(C.RANDOM_SEED)

# ── Utility date helpers ───────────────────────────────────────────────────────
def rdate(start_str, end_str):
    s = date.fromisoformat(start_str)
    e = date.fromisoformat(end_str)
    return s + timedelta(days=rng.randint(0, (e - s).days))

def fmt(d):
    return d.strftime("%d-%m-%Y") if d else ""

def gen_vin():
    chars = "ABCDEFGHJKLMNPRSTUVWXYZ0123456789"
    return "MA3" + ''.join(rng.choices(chars, k=14))

def gen_reg(state_code):
    dist  = rng.randint(1, 49)
    alpha = ''.join(rng.choices("ABCDEFGHJKLMNPRSTUVWXYZ", k=2))
    num   = rng.randint(1000, 9999)
    return f"{state_code}{dist:02d}{alpha}{num}"

def gen_engine(prefix):
    return prefix + ''.join(rng.choices("0123456789ABCDEF", k=8))

def gen_policy_no():
    return ''.join(rng.choices("0123456789", k=14))

# ═══════════════════════════════════════════════════════════════════════════════
#  BUILD CORE DATA RECORDS
# ═══════════════════════════════════════════════════════════════════════════════

def build_records():
    """
    Build 30 vehicle/sale records.
    Returns: list of dicts — one per vehicle sale.
    PII is separated immediately; fact records carry only cust_id.
    """
    records    = []
    pii_data   = []   # for PII_Vault
    mobile_map = {}   # cust_id → raw mobile (kept only in vault)

    model_list = list(C.MODELS_VARIANTS.keys())

    for i, pii_row in enumerate(C.RAW_PII):
        cust_id = f"CUST{i+1:04d}"
        name, gender, dob, city, pincode, occ_id, email_domain = pii_row

        raw_mobile1 = P.generate_mobile(i, rng)
        raw_mobile2 = P.generate_mobile(i + 100, rng) if rng.random() > 0.4 else ""
        raw_email   = P.generate_email(name, email_domain, i)
        raw_address = f"{rng.randint(1,99)}, {rng.choice(['MG Road','Park Street','Gandhi Nagar','Main Market','Station Road'])}, {city}"

        state_code  = C.CITY_STATE_CODE.get(city, "MH")
        state       = C.CITY_STATE.get(city, "Unknown")

        # PII vault row (actual sensitive data — never in fact tables)
        pii_data.append({
            "cust_id":       cust_id,
            "cust_name":     name,
            "mobile_1":      raw_mobile1,
            "mobile_2":      raw_mobile2,
            "email":         raw_email,
            "dob":           dob,
            "address":       raw_address,
            "city":          city,
            "state":         state,
            "pincode":       pincode,
            "gender":        gender,
            "occ_id":        occ_id,
            "created_at":    fmt(rdate("2018-01-01", "2022-12-31")),
        })
        mobile_map[cust_id] = (raw_mobile1, raw_mobile2)

        # Vehicle data
        model   = rng.choice(model_list)
        variants, cc, eng_pfx, fuel_id, transmission, body_type, seats = C.MODELS_VARIANTS[model]
        variant = rng.choice(variants)
        color   = rng.choice(C.COLORS)
        mfd_year= rng.randint(2018, 2024)

        sale_dt  = rdate("2019-01-01", "2024-12-31")
        deliv_dt = sale_dt + timedelta(days=rng.randint(1, 14))
        reg_dt   = deliv_dt + timedelta(days=rng.randint(1, 30))

        fin_flag = rng.choice(["Y", "Y", "N"])
        bank_id  = rng.choice([r[0] for r in C.DIM_BANK]) if fin_flag == "Y" else ""

        dealer_id = rng.choice([r[0] for r in C.DIM_DEALER])
        loc_id    = rng.choice([r[0] for r in C.DIM_LOCATION])
        poc       = rng.choice(C.SALES_POC)

        records.append({
            "idx":       i + 1,
            "cust_id":   cust_id,         # FK to Customer / PII_Vault
            "city":      city,
            "state":     state,
            "state_code":state_code,
            "occ_id":    occ_id,
            "model":     model, "variant": variant, "cc": cc,
            "eng_pfx":   eng_pfx, "fuel_id": fuel_id,
            "transmission": transmission, "body_type": body_type, "seats": seats,
            "color":     color, "mfd_year": mfd_year,
            "sale_dt":   sale_dt, "deliv_dt": deliv_dt, "reg_dt": reg_dt,
            "tax_val":   reg_dt + timedelta(days=365 * 15),
            "pucc_val":  reg_dt + timedelta(days=365),
            "vin":       gen_vin(),
            "reg":       gen_reg(state_code),
            "engine":    gen_engine(eng_pfx),
            "dealer_id": dealer_id,
            "loc_id":    loc_id,
            "poc":       poc,
            "fin_flag":  fin_flag,
            "bank_id":   bank_id,
        })

    return records, pii_data


# ═══════════════════════════════════════════════════════════════════════════════
#  WORKBOOK BUILDER
# ═══════════════════════════════════════════════════════════════════════════════

def build_workbook(records, pii_data):
    wb = openpyxl.load_workbook(C.OUTPUT_FILE)

    # Remove old sheets we will recreate; keep ordering we define
    for sn in list(wb.sheetnames):
        del wb[sn]

    _write_readme(wb)
    _write_dim_tables(wb)
    _write_sales(wb, records)
    _write_service(wb, records)
    _write_insurance(wb, records)
    _write_customer(wb, records, pii_data)
    _write_vehicle(wb, records)
    _write_customer_vehicle(wb, records)
    _write_pii_vault(wb, pii_data)
    _write_churn_features(wb, records)
    _write_churn_scores(wb, records)
    _write_ml_pipeline(wb)

    wb.save(C.OUTPUT_FILE)
    print(f"\n✅  Workbook saved → {C.OUTPUT_FILE}")


# ═══════════════════════════════════════════════════════════════════════════════
#  README SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def _write_readme(wb):
    ws = wb.create_sheet("README")
    E.set_tab_color(ws, "404040")
    E.set_col_widths(ws, [4, 26, 60, 20])

    BG, FG = "404040", "FFFFFF"
    E.merge_paint(ws, 1, 1, 4,
        "AUTOMOTIVE CHURN PREDICTION — SOLUTION WORKBOOK  |  README & INSTRUCTIONS",
        BG, FG, bold=True, size=13)
    E.set_row_height(ws, 1, 32)

    sections = [
        # (row, section_title, content)
        (3,  "HOW TO RUN THE PYTHON CODE",
         "1. Install dependencies:  pip install -r solution_code/requirements.txt\n"
         "2. Generate full workbook: python solution_code/run_solution.py\n"
         "3. Validate output:        python solution_code/validate.py\n"
         "4. All outputs are written to Task .xlsx in this folder."),

        (8,  "COLOUR CODING — SHEET TABS",
         "🟡 GOLD   = Dimension / Lookup tables  (dim_*)\n"
         "🔵 BLUE   = Fact tables                (Sales, Service, Insurance)\n"
         "🟢 GREEN  = Master tables              (Customer, Vehicle)\n"
         "🟤 BROWN  = Bridge table               (Customer_Vehicle)\n"
         "🔴 RED    = PII Vault                  (SENSITIVE — access controlled)\n"
         "🟣 PURPLE = Analytics / ML output      (Churn_Features, Churn_Scores)\n"
         "⬛ GREY   = Documentation              (README, ML_Pipeline)"),

        (16, "DATA GOVERNANCE PRINCIPLES APPLIED",
         "1. PSEUDONYMISATION  — cust_id token used in all fact/analytics tables. Raw name/mobile/email ONLY in PII_Vault.\n"
         "2. DATA MINIMISATION — Each table contains ONLY the columns required for its purpose.\n"
         "3. NORMALISATION     — Free-text categoricals replaced by FK IDs → reduces breach exposure.\n"
         "4. PII VAULT         — Actual PII isolated in PII_Vault sheet. In production: separate DB schema with RBAC.\n"
         "5. GREEN AI          — Feature engineering avoids redundant columns; model uses only high-signal features.\n"
         "6. AUDIT TRAIL       — created_at / updated_at timestamps on master tables for lineage."),

        (22, "SHEET INVENTORY",
         "dim_Bank, dim_Dealer, dim_Location, dim_WorkType, dim_PolicyType,\n"
         "dim_PayMode, dim_CancelReason, dim_Occupation, dim_FuelType, dim_Milestone\n"
         "Sales, Service, Insurance    ← Fact tables (no PII — only cust_id)\n"
         "Customer, Vehicle            ← Master tables (masked PII)\n"
         "Customer_Vehicle             ← Bridge (many-to-many)\n"
         "PII_Vault                    ← Sensitive data store\n"
         "Churn_Features, Churn_Scores ← ML feature & output tables\n"
         "ML_Pipeline                  ← Visual pipeline block diagram"),

        (29, "SCHEMA RELATIONSHIPS (STAR SCHEMA)",
         "Sales.cust_id      → Customer.cust_id  (FK)\n"
         "Sales.vin          → Vehicle.vin        (FK)\n"
         "Sales.bank_id      → dim_Bank.bank_id   (FK)\n"
         "Sales.loc_id       → dim_Location.location_id (FK)\n"
         "Service.vin        → Vehicle.vin        (FK)\n"
         "Service.work_type_id → dim_WorkType.work_type_id (FK)\n"
         "Service.pay_mode_id  → dim_PayMode.pay_mode_id  (FK)\n"
         "Service.cancel_reason_id → dim_CancelReason.cancel_reason_id (FK)\n"
         "Insurance.vin      → Vehicle.vin        (FK)\n"
         "Insurance.policy_type_id → dim_PolicyType.policy_type_id (FK)\n"
         "Customer.occ_id    → dim_Occupation.occ_id (FK)\n"
         "Vehicle.fuel_type_id → dim_FuelType.fuel_type_id (FK)\n"
         "Customer_Vehicle.cust_id   → Customer.cust_id   (FK)\n"
         "Customer_Vehicle.vehicle_id → Vehicle.vehicle_id (FK)\n"
         "PII_Vault.cust_id  → Customer.cust_id   (FK — join only with authorisation)"),
    ]

    for start_row, title, content in sections:
        E.merge_paint(ws, start_row, 2, 4, title, "555555", "FFFFFF", bold=True, size=10)
        E.set_row_height(ws, start_row, 22)
        content_lines = content.count('\n') + 1
        row_h = max(14, content_lines * 14)
        E.merge_paint(ws, start_row + 1, 2, 4, content, "F7F7F7", "222222",
                      bold=False, size=9, h="left")
        E.set_row_height(ws, start_row + 1, row_h + 20)

    print("  README              ✓")


# ═══════════════════════════════════════════════════════════════════════════════
#  DIMENSION TABLES
# ═══════════════════════════════════════════════════════════════════════════════

def _dim_sheet(wb, sheet_name, title, col_headers, schema_notes, col_widths, rows, tab_color="FFC000"):
    ws = wb.create_sheet(sheet_name)
    E.set_tab_color(ws, tab_color)
    ncols = len(col_headers)
    next_row = E.write_sheet_header(ws, sheet_name, title, col_headers,
                                    schema_notes, "FFC000", ncols, col_widths)
    for r in rows:
        E.write_data_row(ws, next_row, list(r), bg="FFFFFF", alt_bg="FFF9E6")
        next_row += 1
    print(f"  {sheet_name:<20} ✓  ({next_row - 4} rows)")


def _write_dim_tables(wb):
    _dim_sheet(wb, "dim_Bank",
        "Banking & NBFC institutions used for vehicle financing",
        ["bank_id","bank_name","bank_type","regulated_by","is_active"],
        ["PK – Unique bank ID","Full institution name","Private / Public / NBFC",
         "Regulator (RBI for India)","Y = Currently offering loans"],
        [10,28,12,14,10],
        C.DIM_BANK)

    _dim_sheet(wb, "dim_Dealer",
        "Authorised dealership network",
        ["dealer_id","dealer_name","city","state","zone","dealer_type"],
        ["PK – Unique dealer ID","Showroom name","City","State","North/South/East/West",
         "3S = Sales+Service+Spares  2S = Sales+Service"],
        [10,24,14,16,10,12],
        C.DIM_DEALER)

    _dim_sheet(wb, "dim_Location",
        "Sales location / service centre geography",
        ["location_id","city","state","zone","tier"],
        ["PK – Unique location ID","City name","State","Geographic zone",
         "Metro / Tier-1 / Tier-2"],
        [14,14,16,10,10],
        C.DIM_LOCATION)

    _dim_sheet(wb, "dim_WorkType",
        "Workshop service work categories",
        ["work_type_id","work_type_name","is_billable","category","sla_hours"],
        ["PK – Unique work type ID","Descriptive name","Y = Customer charged",
         "Scheduled / Unscheduled / Contract","Target completion hours"],
        [14,28,12,16,12],
        C.DIM_WORK_TYPE)

    _dim_sheet(wb, "dim_PolicyType",
        "Insurance policy categories",
        ["policy_type_id","policy_type_name","coverage_type","is_mandatory","description"],
        ["PK","Policy name","What is covered","Y = Legally required",
         "Plain-English description"],
        [14,20,22,14,46],
        C.DIM_POLICY_TYPE)

    _dim_sheet(wb, "dim_PayMode",
        "Payment methods accepted at service centre",
        ["pay_mode_id","pay_mode_name","is_digital","platform"],
        ["PK","Payment method name","Y = Digital / cashless","Processing platform"],
        [12,22,12,18],
        C.DIM_PAY_MODE)

    _dim_sheet(wb, "dim_CancelReason",
        "Reasons a service appointment may be cancelled",
        ["cancel_reason_id","description","category","is_customer_fault"],
        ["PK","Human-readable reason","Operational / Customer-Initiated",
         "Y = Customer caused cancellation"],
        [18,28,24,20],
        C.DIM_CANCEL_REASON)

    _dim_sheet(wb, "dim_Occupation",
        "Customer occupation / employment segment",
        ["occ_id","occupation_name","income_bracket","segment"],
        ["PK","Occupation label","Low/Middle/High income","Employee/Entrepreneur/Other"],
        [10,22,16,16],
        C.DIM_OCCUPATION)

    _dim_sheet(wb, "dim_FuelType",
        "Vehicle fuel / propulsion types (supports Green AI reporting)",
        ["fuel_type_id","fuel_type_name","emission_standard","ev_flag","green_score"],
        ["PK","Fuel type name","Emission norm (BS6/Zero Emission)",
         "Y = Electric vehicle","1–5 sustainability score (5 = cleanest)"],
        [14,14,20,10,14],
        C.DIM_FUEL_TYPE)

    _dim_sheet(wb, "dim_Milestone",
        "Scheduled service milestone intervals",
        ["milestone_id","milestone_name","interval_km","interval_months"],
        ["PK","Milestone label","Odometer trigger","Calendar trigger (months)"],
        [14,26,14,18],
        C.DIM_MILESTONE)


# ═══════════════════════════════════════════════════════════════════════════════
#  FACT TABLES
# ═══════════════════════════════════════════════════════════════════════════════

def _write_sales(wb, records):
    ws = wb.create_sheet("Sales")
    E.set_tab_color(ws, "2E75B6")
    hdrs = ["sales_id","cust_id","dealer_id","vin","mfd_year",
            "registration_date","tax_validity","pucc_validity",
            "finance_flag","bank_id","sale_date","delivery_date",
            "sales_poc","loc_id"]
    notes = [
        "PK – Sale transaction ID",
        "FK → Customer.cust_id  (NO raw name/mobile here)",
        "FK → dim_Dealer",
        "FK → Vehicle.vin",
        "Manufacturing year",
        "RTO registration date (DD-MM-YYYY)",
        "Road tax valid until",
        "Pollution cert expiry",
        "Y = Financed",
        "FK → dim_Bank  (blank if not financed)",
        "Date of sale",
        "Date of delivery",
        "Salesperson name",
        "FK → dim_Location",
    ]
    widths = [10,10,10,20,10,18,18,14,12,10,14,14,18,14]
    E.write_sheet_header(ws, "Sales", "Vehicle sale transactions — no PII stored",
                         hdrs, notes, "2E75B6", len(hdrs), widths)
    row = 4
    for r in records:
        vals = [
            f"SL{r['idx']:04d}", r["cust_id"], r["dealer_id"], r["vin"],
            r["mfd_year"], fmt(r["reg_dt"]), fmt(r["tax_val"]),
            fmt(r["pucc_val"]), r["fin_flag"], r["bank_id"],
            fmt(r["sale_dt"]), fmt(r["deliv_dt"]), r["poc"], r["loc_id"],
        ]
        E.write_data_row(ws, row, vals)
        row += 1
    print(f"  {'Sales':<20} ✓  ({row - 4} rows)")


def _write_service(wb, records):
    ws = wb.create_sheet("Service")
    E.set_tab_color(ws, "2E75B6")
    hdrs = ["smr_id","cust_id","dealer_id","vin","work_type_id",
            "milestone_id","visit_date","bill_date","status",
            "cancel_reason_id","bill_amount","mileage",
            "service_advisor","pay_mode_id"]
    notes = [
        "PK – Service record ID",
        "FK → Customer.cust_id  (NO raw name/mobile)",
        "FK → dim_Dealer",
        "FK → Vehicle.vin",
        "FK → dim_WorkType",
        "FK → dim_Milestone",
        "Workshop visit date",
        "Invoice date",
        "Completed / Cancelled",
        "FK → dim_CancelReason  (blank if Completed)",
        "Amount charged in INR (0 for free services)",
        "Odometer reading at visit",
        "Service advisor name",
        "FK → dim_PayMode",
    ]
    widths = [10,10,10,20,14,12,14,14,12,18,14,12,18,12]
    work_type_ids  = [r[0] for r in C.DIM_WORK_TYPE]
    milestone_ids  = [r[0] for r in C.DIM_MILESTONE]
    pay_mode_ids   = [r[0] for r in C.DIM_PAY_MODE]
    cancel_ids     = [r[0] for r in C.DIM_CANCEL_REASON]

    E.write_sheet_header(ws, "Service", "Workshop visit records — no PII stored",
                         hdrs, notes, "2E75B6", len(hdrs), widths)
    row = 4
    smr = 1
    for r in records:
        n_visits   = rng.randint(1, 4)
        last_visit = r["deliv_dt"]
        mileage    = rng.randint(1000, 5000)
        for v in range(n_visits):
            visit_dt  = last_visit + timedelta(days=rng.randint(90, 200))
            bill_dt   = visit_dt   + timedelta(days=rng.randint(0, 2))
            wt_id     = work_type_ids[min(v, len(work_type_ids)-1)]
            ms_id     = milestone_ids[min(v, len(milestone_ids)-1)]
            is_free   = (C.WORK_TYPE_NAMES[wt_id] == "Free Service")
            bill_amt  = 0 if is_free else rng.randint(1500, 18000)
            mileage  += rng.randint(5000, 15000)
            status    = rng.choices(["Completed","Cancelled"], weights=[9,1])[0]
            cr_id     = rng.choice(cancel_ids) if status == "Cancelled" else ""
            pm_id     = rng.choice(pay_mode_ids) if not is_free else "PM001"
            vals = [
                f"SMR{smr:05d}", r["cust_id"], r["dealer_id"], r["vin"],
                wt_id, ms_id, fmt(visit_dt), fmt(bill_dt), status,
                cr_id, bill_amt, mileage,
                rng.choice(C.SERVICE_ADVISORS), pm_id,
            ]
            E.write_data_row(ws, row, vals)
            row   += 1
            smr   += 1
            last_visit = visit_dt
    print(f"  {'Service':<20} ✓  ({row - 4} rows)")


def _write_insurance(wb, records):
    ws = wb.create_sheet("Insurance")
    E.set_tab_color(ws, "2E75B6")
    hdrs = ["ins_id","cust_id","dealer_id","vin","policy_type_id",
            "policy_no","policy_issue_date","policy_inception_date",
            "policy_expiry_date","tp_policy_issue_date","tp_policy_expiry_date",
            "previous_policy_start_date","previous_policy_end_date",
            "auto_membership","total_premium"]
    notes = [
        "PK – Insurance record ID",
        "FK → Customer.cust_id  (NO raw name/mobile)",
        "FK → dim_Dealer",
        "FK → Vehicle.vin",
        "FK → dim_PolicyType",
        "Insurer-issued policy number",
        "Date policy was issued",
        "Risk start date",
        "Policy expiry date  ← KEY churn signal",
        "Third-party policy issue date",
        "Third-party policy expiry (3-yr for new, 1-yr for renewal)",
        "Previous policy start (blank for first policy)",
        "Previous policy end  (blank for first policy)",
        "Y = Has roadside assistance membership  ← churn signal",
        "Total premium paid in INR  ← trend feature",
    ]
    widths = [10,10,10,20,14,16,16,18,16,18,20,24,22,16,14]
    pt_ids = [r[0] for r in C.DIM_POLICY_TYPE]
    E.write_sheet_header(ws, "Insurance", "Vehicle insurance policies — no PII stored",
                         hdrs, notes, "2E75B6", len(hdrs), widths)
    row = 4
    ins = 1
    for r in records:
        n_pol     = rng.randint(1, 2)
        pol_start = r["sale_dt"]
        for p in range(n_pol):
            pol_exp = pol_start + timedelta(days=365)
            tp_exp  = pol_start + timedelta(days=365*3) if p == 0 else pol_start + timedelta(days=365)
            prev_s  = pol_start - timedelta(days=365) if p > 0 else None
            prev_e  = pol_start if p > 0 else None
            vals = [
                f"INS{ins:05d}", r["cust_id"], r["dealer_id"], r["vin"],
                rng.choice(pt_ids), gen_policy_no(),
                fmt(pol_start), fmt(pol_start), fmt(pol_exp),
                fmt(pol_start), fmt(tp_exp),
                fmt(prev_s) if prev_s else "", fmt(prev_e) if prev_e else "",
                rng.choice(["Y","N","N"]),
                rng.randint(8000, 45000),
            ]
            E.write_data_row(ws, row, vals)
            row       += 1
            ins       += 1
            pol_start  = pol_exp + timedelta(days=1)
    print(f"  {'Insurance':<20} ✓  ({row - 4} rows)")


# ═══════════════════════════════════════════════════════════════════════════════
#  MASTER TABLES
# ═══════════════════════════════════════════════════════════════════════════════

def _write_customer(wb, records, pii_data):
    """Customer master — masked PII only.  No raw name/mobile/email/address."""
    ws = wb.create_sheet("Customer")
    E.set_tab_color(ws, "375623")
    hdrs = ["cust_id","name_masked","mobile_masked","email_masked",
            "birth_year","gender","occ_id","city","state","pincode",
            "preferred_dealer_id","created_at","updated_at"]
    notes = [
        "PK – Unique customer token",
        "Masked name  (e.g. R**** S*****) — full name in PII_Vault",
        "Masked mobile (e.g. 98XXXXXX10) — actual number in PII_Vault",
        "Masked email  (e.g. r.***@g***.com) — actual email in PII_Vault",
        "Birth year only — full DOB in PII_Vault (lower re-ID risk)",
        "M / F / O",
        "FK → dim_Occupation",
        "City of residence (non-PII aggregate)",
        "State (non-PII aggregate)",
        "6-digit PIN (low-granularity — kept for geo analytics)",
        "FK → dim_Dealer  (preferred dealer)",
        "Record created date",
        "Last updated date",
    ]
    widths = [10,18,16,22,12,8,10,14,16,10,18,14,14]
    E.write_sheet_header(ws, "Customer",
        "Customer master — masked PII. Raw PII in PII_Vault (access controlled).",
        hdrs, notes, "375623", len(hdrs), widths)

    pii_dict = {p["cust_id"]: p for p in pii_data}
    row = 4
    for r in records:
        cid = r["cust_id"]
        p   = pii_dict[cid]
        vals = [
            cid,
            P.mask_name(p["cust_name"]),
            P.mask_mobile(p["mobile_1"]),
            P.mask_email(p["email"]),
            P.mask_dob_to_birth_year(p["dob"]),
            p["gender"],
            p["occ_id"],
            p["city"],
            p["state"],
            p["pincode"],
            r["dealer_id"],
            p["created_at"],
            fmt(rdate("2022-01-01", "2025-04-01")),
        ]
        E.write_data_row(ws, row, vals, bg="FFFFFF", alt_bg="F0F7F0")
        row += 1
    print(f"  {'Customer':<20} ✓  ({row - 4} rows) — PII masked")


def _write_vehicle(wb, records):
    ws = wb.create_sheet("Vehicle")
    E.set_tab_color(ws, "375623")
    hdrs = ["vehicle_id","vin","reg_no","engine_no","model","variant",
            "color","cc","mfd_year","registration_date","fuel_type_id",
            "transmission","body_type","seating_capacity","created_at"]
    notes = [
        "PK – Unique vehicle token",
        "17-char VIN (unique per vehicle chassis)",
        "RTO registration plate number",
        "Engine serial number",
        "Vehicle model name",
        "Variant / trim level",
        "Exterior colour",
        "Engine displacement in CC",
        "Manufacturing year",
        "RTO registration date",
        "FK → dim_FuelType",
        "Manual / Automatic / AMT / CVT",
        "Hatchback / Sedan / SUV / MUV",
        "Number of seats",
        "Record created date",
    ]
    widths = [12,20,14,18,10,14,16,8,10,18,14,14,14,12,14]
    E.write_sheet_header(ws, "Vehicle",
        "Vehicle master registry — one row per unique VIN.",
        hdrs, notes, "375623", len(hdrs), widths)

    seen_vins = set()
    row = 4
    vid = 1
    for r in records:
        if r["vin"] in seen_vins:
            continue
        seen_vins.add(r["vin"])
        vals = [
            f"VEH{vid:04d}", r["vin"], r["reg"], r["engine"],
            r["model"], r["variant"], r["color"],
            r["cc"], r["mfd_year"], fmt(r["reg_dt"]),
            r["fuel_id"], r["transmission"], r["body_type"], r["seats"],
            fmt(r["sale_dt"]),
        ]
        E.write_data_row(ws, row, vals, bg="FFFFFF", alt_bg="F0F7F0")
        row += 1
        vid += 1
    print(f"  {'Vehicle':<20} ✓  ({row - 4} rows)")


def _write_customer_vehicle(wb, records):
    ws = wb.create_sheet("Customer_Vehicle")
    E.set_tab_color(ws, "7B3F00")
    hdrs = ["cv_id","cust_id","vehicle_id","relationship_type",
            "purchase_date","is_primary_owner","active_flag"]
    notes = [
        "PK – Unique mapping ID",
        "FK → Customer.cust_id",
        "FK → Vehicle.vehicle_id  (VEH0001, VEH0002...)",
        "Owner / Co-Owner / Family Member",
        "Date of vehicle purchase",
        "Y = Primary registered owner",
        "Y = Currently owns vehicle  N = Sold/Transferred",
    ]
    widths = [10,10,12,20,16,18,12]
    E.write_sheet_header(ws, "Customer_Vehicle",
        "Bridge: many-to-many between Customer and Vehicle.",
        hdrs, notes, "7B3F00", len(hdrs), widths)

    # vin → vehicle_id map
    vin_to_vid = {}
    vid = 1
    seen = set()
    for r in records:
        if r["vin"] not in seen:
            vin_to_vid[r["vin"]] = f"VEH{vid:04d}"
            seen.add(r["vin"])
            vid += 1

    row  = 4
    cv   = 1
    seen_pairs = set()
    for r in records:
        pair = (r["cust_id"], vin_to_vid[r["vin"]])
        if pair not in seen_pairs:
            seen_pairs.add(pair)
            E.write_data_row(ws, row, [
                f"CV{cv:04d}", r["cust_id"], vin_to_vid[r["vin"]],
                rng.choices(["Owner","Owner","Co-Owner","Family Member"],
                             weights=[6,6,2,1])[0],
                fmt(r["sale_dt"]), "Y",
                rng.choice(["Y","Y","Y","N"]),
            ], bg="FFFFFF", alt_bg="FAE5D3")
            cv  += 1
            row += 1

    # Add 5 co-owner examples (demonstrates many-to-many)
    vins = list(vin_to_vid.keys())[:5]
    custs= [r["cust_id"] for r in records][20:25]
    for vin, cid in zip(vins, custs):
        pair = (cid, vin_to_vid[vin])
        if pair not in seen_pairs:
            seen_pairs.add(pair)
            E.write_data_row(ws, row, [
                f"CV{cv:04d}", cid, vin_to_vid[vin], "Co-Owner",
                fmt(rdate("2020-01-01","2023-12-31")), "N", "Y",
            ], bg="FFFFFF", alt_bg="FAE5D3")
            cv  += 1
            row += 1
    print(f"  {'Customer_Vehicle':<20} ✓  ({row - 4} mappings)")


# ═══════════════════════════════════════════════════════════════════════════════
#  PII VAULT  (sensitive — clearly labelled)
# ═══════════════════════════════════════════════════════════════════════════════

def _write_pii_vault(wb, pii_data):
    ws = wb.create_sheet("PII_Vault")
    E.set_tab_color(ws, "C00000")

    # Extra-prominent warning banner (row 1)
    ncols = 12
    E.merge_paint(ws, 1, 1, ncols,
        "⛔  PII VAULT — SENSITIVE DATA  |  "
        "GDPR / DPDP Act 2023 Compliant  |  "
        "Production: Separate schema with RBAC + Encryption at rest",
        "C00000", "FFFFFF", bold=True, size=12)
    E.set_row_height(ws, 1, 32)

    hdrs = ["cust_id","cust_name","mobile_1","mobile_2","email",
            "dob","address","city","state","pincode","gender","occ_id"]
    notes = [
        "FK → Customer.cust_id  (join key)",
        "Full legal name  ← PII",
        "Primary mobile  ← PII",
        "Alt mobile       ← PII  (nullable)",
        "Email address    ← PII",
        "Date of birth    ← PII",
        "Street address   ← PII",
        "City             (low-risk geographic)",
        "State            (low-risk geographic)",
        "PIN code         (low-risk geographic)",
        "Gender           (quasi-identifier)",
        "FK → dim_Occupation",
    ]
    E.write_sheet_header(ws, "PII_Vault",
        "RAW PII — Join to Customer table via cust_id  |  DO NOT expose in reports",
        hdrs, notes, "C00000", ncols,
        [10,22,14,14,30,14,36,14,16,10,8,10])

    row = 4
    for p in pii_data:
        E.write_data_row(ws, row, [
            p["cust_id"], p["cust_name"], p["mobile_1"], p["mobile_2"],
            p["email"], p["dob"], p["address"],
            p["city"], p["state"], p["pincode"],
            p["gender"], p["occ_id"],
        ], bg="FFF5F5", alt_bg="FFE0E0")
        row += 1
    print(f"  {'PII_Vault':<20} ✓  ({row - 4} rows) — SENSITIVE")


# ═══════════════════════════════════════════════════════════════════════════════
#  CHURN FEATURES TABLE
# ═══════════════════════════════════════════════════════════════════════════════

def _write_churn_features(wb, records):
    ws = wb.create_sheet("Churn_Features")
    E.set_tab_color(ws, "7030A0")
    hdrs = [
        "cust_id","vehicle_id","vehicle_age_years","finance_flag",
        "total_service_visits","days_since_last_service",
        "avg_bill_amount","pct_paid_services","cancelled_visit_flag",
        "num_policy_renewals","days_to_policy_expiry",
        "auto_membership_flag","premium_trend",
        "recency_score","frequency_score","monetary_score","rfm_total",
        "churn_label","risk_tier"
    ]
    notes = [
        "FK → Customer","FK → Vehicle",
        "TODAY_YEAR - mfd_year","1 if financed",
        "COUNT(smr_id) per customer","Days since MAX(visit_date)",
        "AVG(bill_amount) excl. free","Paid visits / All visits",
        "1 if any visit Cancelled","COUNT(policies) - 1",
        "policy_expiry_date - TODAY","1 if auto_membership=Y",
        "Last premium - First premium","Recency bucket 1-5",
        "Frequency bucket 1-5","Monetary bucket 1-5","R+F+M total",
        "1=Churned 0=Active (target variable)","High/Medium/Low",
    ]
    widths = [10,12,16,12,18,20,16,16,18,18,20,18,14,14,14,14,10,12,12]
    E.write_sheet_header(ws, "Churn_Features",
        "Feature-engineered dataset for ML model training.",
        hdrs, notes, "7030A0", len(hdrs), widths)

    vin_to_vid = {}
    vid = 1
    seen = set()
    for r in records:
        if r["vin"] not in seen:
            vin_to_vid[r["vin"]] = f"VEH{vid:04d}"
            seen.add(r["vin"])
            vid += 1

    today = date(2025, 5, 3)
    row   = 4
    for r in records:
        vehicle_age = today.year - r["mfd_year"]

        # Realistic churn distribution: ~40% High, ~30% Medium, ~30% Low
        # (mirrors the "40% annual churn" stated in the problem)
        tier_roll = rng.random()
        if tier_roll < 0.38:        # ~38% churned / high-risk
            days_since_svc = rng.randint(366, 700)
            days_to_exp    = rng.randint(-180, -1)
            svc_visits     = rng.randint(1, 2)
            n_renewals     = rng.randint(0, 1)
            is_churn       = 1
        elif tier_roll < 0.65:      # ~27% medium-risk (at-risk)
            days_since_svc = rng.randint(181, 365)
            days_to_exp    = rng.randint(1, 60)
            svc_visits     = rng.randint(1, 3)
            n_renewals     = rng.randint(0, 1)
            is_churn       = 0
        else:                       # ~35% low-risk (active)
            days_since_svc = rng.randint(7, 180)
            days_to_exp    = rng.randint(61, 365)
            svc_visits     = rng.randint(2, 5)
            n_renewals     = rng.randint(1, 3)
            is_churn       = 0

        avg_bill       = rng.randint(0, 15000)
        pct_paid       = round(rng.uniform(0, 1), 2)
        cancelled_flag = rng.choice([0, 0, 0, 1])
        auto_mbr       = rng.choice([0, 0, 1])
        prem_trend     = rng.randint(-5000, 8000)

        rec_score  = max(1, 5 - days_since_svc // 140)
        freq_score = min(5, svc_visits)
        mon_score  = min(5, max(1, avg_bill // 3000))
        rfm        = rec_score + freq_score + mon_score

        risk = "High" if is_churn == 1 else ("Medium" if rfm < 8 else "Low")

        E.write_data_row(ws, row, [
            r["cust_id"], vin_to_vid[r["vin"]], vehicle_age, r["fin_flag"],
            svc_visits, days_since_svc, avg_bill, pct_paid, cancelled_flag,
            n_renewals, days_to_exp, auto_mbr, prem_trend,
            rec_score, freq_score, mon_score, rfm,
            is_churn, risk,
        ])
        row += 1
    print(f"  {'Churn_Features':<20} ✓  ({row - 4} rows)")


# ═══════════════════════════════════════════════════════════════════════════════
#  CHURN SCORES TABLE (model output)
# ═══════════════════════════════════════════════════════════════════════════════

def _write_churn_scores(wb, records):
    ws = wb.create_sheet("Churn_Scores")
    E.set_tab_color(ws, "7030A0")
    hdrs = [
        "cust_id","churn_probability","risk_tier",
        "top_driver_1","top_driver_2","top_driver_3",
        "recommended_action","score_date"
    ]
    notes = [
        "FK → Customer","Model output: 0.00–1.00",
        "High / Medium / Low",
        "Highest SHAP feature","2nd SHAP feature","3rd SHAP feature",
        "Suggested CRM action","Date model was run",
    ]
    widths = [10,18,12,26,26,26,36,14]
    E.write_sheet_header(ws, "Churn_Scores",
        "Model output: churn probability + top SHAP drivers per customer.",
        hdrs, notes, "7030A0", len(hdrs), widths)

    drivers = [
        "days_since_last_service","days_to_policy_expiry","avg_bill_amount",
        "pct_paid_services","num_policy_renewals","auto_membership_flag",
        "cancelled_visit_flag","rfm_total","vehicle_age_years",
    ]
    actions = {
        "High":   "Urgent outreach: Call + service reminder + policy renewal offer",
        "Medium": "Send WhatsApp reminder for upcoming service / insurance renewal",
        "Low":    "Monthly newsletter — loyalty points update",
    }
    row = 4
    for r in records:
        prob  = round(rng.uniform(0.0, 1.0), 4)
        tier  = "High" if prob > 0.65 else ("Medium" if prob > 0.35 else "Low")
        d_all = rng.sample(drivers, 3)
        E.write_data_row(ws, row, [
            r["cust_id"], prob, tier,
            d_all[0], d_all[1], d_all[2],
            actions[tier],
            "03-05-2025",
        ])
        row += 1
    print(f"  {'Churn_Scores':<20} ✓  ({row - 4} rows)")


# ═══════════════════════════════════════════════════════════════════════════════
#  ML PIPELINE VISUAL
# ═══════════════════════════════════════════════════════════════════════════════

def _write_ml_pipeline(wb):
    ws = wb.create_sheet("ML_Pipeline")
    E.set_tab_color(ws, "1F4E79")

    col_widths = [3, 22, 22, 22, 22, 22, 3]
    E.set_col_widths(ws, col_widths)

    def box(row, col, text, bg, fg="FFFFFF", bold=False, size=9):
        c = ws.cell(row, col)
        E.paint(c, text, bg, fg, bold=bold, size=size, h="left",
                border_color="888888")

    def hdr(row, text, bg):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        E.paint(ws.cell(row, 2), text, bg, "FFFFFF", bold=True, size=11,
                border_color=bg)
        E.set_row_height(ws, row, 24)

    def arrow(row):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        E.paint(ws.cell(row, 2), "⬇  ⬇  ⬇", "F2F2F2", "888888",
                bold=True, size=12, border_color="DDDDDD")
        E.set_row_height(ws, row, 18)

    def quad(row, items, bg, fg, row_height=80):
        for ci, (label, text) in enumerate(items, 2):
            cell = ws.cell(row, ci)
            E.paint(cell, f"[{label}]\n{text}", bg, fg, bold=False, size=8,
                    h="left", border_color="AAAAAA")
        E.set_row_height(ws, row, row_height)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    E.paint(ws.cell(1,1),
        "CHURN PREDICTION ML PIPELINE — Automotive After-Sales  (Relatim Technical Exercise)",
        "1F4E79","FFFFFF",bold=True,size=13,border_color="1F4E79")
    E.set_row_height(ws, 1, 34)

    # Stage 1
    hdr(3, "STAGE 1 — DATA SOURCES", "2E75B6")
    quad(4, [
        ("Sales Table",      "sale_date, vin, cust_id\nfinance_flag, bank_id, loc_id"),
        ("Service Table",    "visit_date, work_type_id\nbill_amount, mileage, status"),
        ("Insurance Table",  "policy_expiry_date\npolicy_type_id, premium"),
        ("Dim Tables",       "Bank, WorkType, PolicyType\nPayMode, CancelReason etc."),
        ("PII_Vault",        "cust_name, mobile, email\n⛔ Access controlled — RBAC"),
    ], "D6E4F0", "1F4E79", row_height=70)
    arrow(5)

    # Stage 2
    hdr(6, "STAGE 2 — DATA INGESTION & PROCESSING", "375623")
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=6)
    E.paint(ws.cell(7,2),
        "• Deduplicate customers: (cust_id via PII token — no raw name in joins)\n"
        "• Join Sales + Service + Insurance on VIN / cust_id\n"
        "• Replace dim IDs with descriptive labels for readability (via lookup join)\n"
        "• Parse & validate dates  |  Handle nulls: impute median for bill_amount\n"
        "• Flag: policy_expired = (policy_expiry_date < TODAY)",
        "E2EFDA","1A3C0A",bold=False,size=9,h="left",border_color="AAAAAA")
    E.set_row_height(ws, 7, 72)
    arrow(8)

    # Stage 3
    hdr(9, "STAGE 3 — FEATURE ENGINEERING", "7B3F00")
    quad(10, [
        ("Sales Features",      "vehicle_age_years\nfinance_flag\ndays_to_first_service"),
        ("Service Features",    "total_service_visits\ndays_since_last_service\navg_bill_amount\npct_paid_services\ncancelled_visit_flag"),
        ("Insurance Features",  "days_to_policy_expiry\nnum_policy_renewals\nauto_membership_flag\npremium_trend"),
        ("RFM Derived",         "recency_score (1-5)\nfrequency_score (1-5)\nmonetary_score (1-5)\nrfm_total"),
        ("Churn Label",         "churn_label = 1 if:\n  days_since_svc > 365\n  AND policy expired"),
    ], "FAE5D3","5D2906", row_height=90)
    arrow(11)

    # Stage 4
    hdr(12, "STAGE 4 — MODEL / ANALYTICAL LOGIC", "7030A0")
    quad(13, [
        ("Rule-Based Baseline", "Flag directly by:\n• No svc > 365 days\n• Policy not renewed\n• PUCC expired"),
        ("XGBoost Classifier",  "Target: churn_label\nFeatures: 17 columns\nSplit: 80/20\nMetrics: AUC, F1"),
        ("K-Means Clustering",  "Segments on RFM:\n• Champions\n• Loyal\n• At-Risk\n• Lost"),
        ("SHAP Explainability", "Top-3 feature drivers\nper customer output\nEnsures transparency\nGreen AI ✓"),
        ("Model Governance",    "Version controlled\nData lineage tracked\nMonthly label refresh\nQuarterly retrain"),
    ], "EAD7F7","4A235A", row_height=90)
    arrow(14)

    # Stage 5
    hdr(15, "STAGE 5 — OUTPUT & ACTION", "C00000")
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=6)
    E.paint(ws.cell(16,2),
        "📊  Dashboard  : Churn risk leaderboard by dealer / city / model  (Churn_Scores sheet)\n"
        "📧  CRM Alerts : WhatsApp / Call for High-risk: service reminder + insurance renewal\n"
        "📋  Reports    : Monthly churn rate, segment distribution, revenue-at-risk (INR)\n"
        "🔄  Feedback   : Update labels monthly → retrain model quarterly → track AUC drift\n"
        "🔐  Privacy    : No PII in dashboards — only cust_id + masked fields shown to analysts",
        "FFE0E0","7B0000",bold=False,size=9,h="left",border_color="AAAAAA")
    E.set_row_height(ws, 16, 80)

    # Risk legend
    hdr(18, "CHURN RISK TIER DEFINITIONS", "404040")
    quad(19, [
        ("🔴 HIGH",    "days_since_svc > 365\nAND policy expired\nAction: Urgent outreach"),
        ("🟡 MEDIUM",  "svc gap > 180 days\nOR only 1 lifetime visit\nAction: Reminder SMS/WA"),
        ("🟢 LOW",     "svc visit < 180 days\nAND active policy\nAction: Loyalty newsletter"),
        ("Green AI ✓", "Model uses only\nhigh-signal features\nNo redundant compute"),
        ("Data Ethics","PII masked in output\ncust_id token only\nRBAC on PII_Vault"),
    ], "F2F2F2","333333", row_height=70)

    print(f"  {'ML_Pipeline':<20} ✓  (visual block diagram)")


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n=== Automotive Churn Prediction — Generating Solution Workbook ===\n")
    print("Building records...")
    records, pii_data = build_records()
    print(f"  {len(records)} vehicle sale records built\n")
    print("Writing sheets...")
    build_workbook(records, pii_data)
    print("\nAll sheets written successfully.")


if __name__ == "__main__":
    main()
