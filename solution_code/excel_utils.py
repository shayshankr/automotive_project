"""
excel_utils.py
==============
Reusable Excel formatting helpers (openpyxl).

Colour Coding Convention (shown in README sheet):
  Dark Blue  (#1F4E79) — Fact tables   (Sales, Service, Insurance)
  Dark Green (#375623) — Master tables (Customer, Vehicle)
  Gold/Amber (#FFC000) — Dimension tables (dim_*)
  Red        (#C00000) — PII Vault (sensitive — access controlled)
  Brown      (#7B3F00) — Bridge tables (Customer_Vehicle)
  Purple     (#7030A0) — Analytics / ML output
  Dark Grey  (#404040) — README / Documentation
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter


# ── Border presets ────────────────────────────────────────────────────────────
def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border(color="888888"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Core cell painter ─────────────────────────────────────────────────────────
def paint(cell, value, bg, fg="FFFFFF", bold=False, size=10,
          h="center", v="center", wrap=True, italic=False, border_color="CCCCCC"):
    cell.value     = value
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(name="Arial", color=fg, bold=bold, size=size, italic=italic)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.border    = thin_border(border_color)
    return cell


def merge_paint(ws, row, c1, c2, value, bg, fg="FFFFFF",
                bold=False, size=10, h="center", italic=False):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1)
    paint(c, value, bg, fg, bold=bold, size=size, h=h, italic=italic,
          border_color="555555")
    return c


# ── Sheet-level helpers ───────────────────────────────────────────────────────
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def freeze(ws, cell="A4"):
    ws.freeze_panes = cell

def set_tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color

def clear_sheet(ws):
    """Delete every row (including header/title rows)."""
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row + 20)


# ── Standard sheet header block (3 rows) ──────────────────────────────────────
def write_sheet_header(ws, title, subtitle, col_headers, schema_notes,
                       header_bg, ncols, col_widths=None):
    """
    Row 1 : Merged title banner
    Row 2 : Column headers
    Row 3 : Schema notes (italic, light background)
    Returns next available row number (4).
    """
    light_bg = _lighten(header_bg)

    # Row 1 — title
    merge_paint(ws, 1, 1, ncols, f"{title}  |  {subtitle}",
                header_bg, "FFFFFF", bold=True, size=12)
    set_row_height(ws, 1, 28)

    # Row 2 — column headers
    for ci, h in enumerate(col_headers, 1):
        paint(ws.cell(2, ci), h, header_bg, "FFFFFF", bold=True, size=10,
              border_color=header_bg)
    set_row_height(ws, 2, 20)

    # Row 3 — schema notes
    for ci, note in enumerate(schema_notes, 1):
        paint(ws.cell(3, ci), note, light_bg, header_bg,
              bold=False, size=9, italic=True, h="left", border_color="DDDDDD")
    set_row_height(ws, 3, 32)

    if col_widths:
        set_col_widths(ws, col_widths)

    freeze(ws, "A4")
    return 4   # next data row


def write_data_row(ws, row_idx, values, bg="FFFFFF", fg="222222",
                   bold=False, alt_bg="F7F7F7"):
    """Write a single data row; alternating row shading."""
    fill_bg = alt_bg if row_idx % 2 == 0 else bg
    for ci, v in enumerate(values, 1):
        paint(ws.cell(row_idx, ci), v, fill_bg, fg,
              bold=bold, size=10, h="left", border_color="DDDDDD")


# ── Private helpers ───────────────────────────────────────────────────────────
_LIGHTEN_MAP = {
    "1F4E79": "D6E4F0",  # dark blue  → light blue
    "375623": "E2EFDA",  # dark green → light green
    "FFC000": "FFF2CC",  # gold       → light gold
    "C00000": "FFE0E0",  # red        → light red
    "7B3F00": "FAE5D3",  # brown      → light brown
    "7030A0": "EAD7F7",  # purple     → light purple
    "404040": "F2F2F2",  # grey       → light grey
}
def _lighten(hex_color):
    return _LIGHTEN_MAP.get(hex_color, "F2F2F2")
